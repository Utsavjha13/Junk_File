import openpyxl
from openpyxl.styles import PatternFill


def get_max_row(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def get_max_col(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def read_data(file,sheetName,rowNum,colNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    val=sheet.cell(rowNum,colNum).value
    return val

def write_data(file,sheetName,rowNum,colNum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rowNum, colNum).value=data
    workbook.save(file)

def green_color_cell(file,sheetName,rowNum,colNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenfill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rowNum,colNum).fill=greenfill
    workbook.save(file)

def red_color_cell(file,sheetName,rowNum,colNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    redfill=PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rowNum,colNum).fill=redfill
    workbook.save(file)
