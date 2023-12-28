import time
import random
import openpyxl

file="/Users/utsav.jha/Downloads/moneyControl.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
Period=["day(s)","month(s)","year(s)"]
Freq=["Simple Interest","Compounded Monthly","Compounded Quarterly","Compounded Half Yearly"]
for i in range (2,31):
    sheet.cell(i,1).value= random.randrange(100000,5000000,50000)
    sheet.cell(i,2).value=random.randint(6,13)
    sheet.cell(i,3).value=random.randrange(5,25,1)
    sheet.cell(i,4).value=random.choice(Period)
    sheet.cell(i,5).value=random.choice(Freq)
    workbook.save(file)




