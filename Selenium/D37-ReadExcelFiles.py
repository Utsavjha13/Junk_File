import time
import openpyxl
from selenium import webdriver

file= "/Users/utsav.jha/Downloads/Device_GUID.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet3"]
rows=sheet.max_row
col=sheet.max_column

for i in range (1,rows+1):
    for j in range(1, col+1):
        print(sheet.cell(i,j).value, end="                            ")

    print()